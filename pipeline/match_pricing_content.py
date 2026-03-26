from __future__ import annotations

import argparse
import csv
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_INPUT_DIR = Path(__file__).resolve().parent / "input"
DEFAULT_CONTENT_CSV = DEFAULT_INPUT_DIR / "contentLicensingnew.csv"
DEFAULT_PRICING_XLSX = DEFAULT_INPUT_DIR / "pricing.xlsx"
MODEL_COLUMN = "Model No./No modèle"
CONTENT_PROD_COLUMN = "Prod"
WAREHOUSE_COLUMN = "warehouse_location"


def normalize_key(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def read_content_prod_values(csv_path: Path) -> set[str]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        if not reader.fieldnames or CONTENT_PROD_COLUMN not in reader.fieldnames:
            raise ValueError(
                f"Missing '{CONTENT_PROD_COLUMN}' column in {csv_path}. "
                f"Found columns: {reader.fieldnames}"
            )

        values = {
            normalize_key(row.get(CONTENT_PROD_COLUMN))
            for row in reader
            if normalize_key(row.get(CONTENT_PROD_COLUMN))
        }
    return values


def _sheet_rows_with_headers(worksheet) -> Iterable[dict[str, object]]:
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        return

    for row in rows:
        data = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        data[WAREHOUSE_COLUMN] = worksheet.title
        yield data


def count_matches(pricing_path: Path, content_prod_values: set[str]) -> tuple[int, int]:
    workbook = load_workbook(pricing_path, data_only=True, read_only=True)
    total_rows = 0
    matched_rows = 0

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]

        # Empty sheets are ignored.
        rows_iter = _sheet_rows_with_headers(worksheet)
        for row in rows_iter:
            if MODEL_COLUMN not in row:
                raise ValueError(
                    f"Missing '{MODEL_COLUMN}' column in sheet '{sheet_name}' "
                    f"of {pricing_path}. Available columns: {list(row.keys())}"
                )

            total_rows += 1
            model_no = normalize_key(row.get(MODEL_COLUMN))
            if model_no and model_no in content_prod_values:
                matched_rows += 1

    return total_rows, matched_rows


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Count how many pricing rows match content licensing records by "
            "Model No./No modèle -> Prod."
        )
    )
    parser.add_argument(
        "--content-csv",
        default=str(DEFAULT_CONTENT_CSV),
        help=f"Path to content licensing CSV. Default: {DEFAULT_CONTENT_CSV}",
    )
    parser.add_argument(
        "--pricing-xlsx",
        default=str(DEFAULT_PRICING_XLSX),
        help=f"Path to pricing workbook. Default: {DEFAULT_PRICING_XLSX}",
    )

    args = parser.parse_args()
    content_csv = Path(args.content_csv)
    pricing_xlsx = Path(args.pricing_xlsx)

    if not content_csv.exists():
        raise FileNotFoundError(f"Content CSV not found: {content_csv}")
    if not pricing_xlsx.exists():
        raise FileNotFoundError(f"Pricing workbook not found: {pricing_xlsx}")

    prod_values = read_content_prod_values(content_csv)
    total_rows, matched_rows = count_matches(pricing_xlsx, prod_values)

    print(f"Total pricing rows processed: {total_rows}")
    print(f"Matching rows found: {matched_rows}")


if __name__ == "__main__":
    main()

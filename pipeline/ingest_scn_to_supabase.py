from __future__ import annotations

import argparse
import asyncio
import csv
import logging
import os
from pathlib import Path
import re
import sys
from typing import Iterable

from openpyxl import load_workbook

API_ROOT = Path(__file__).resolve().parents[1] / "api"
if str(API_ROOT) not in sys.path:
    sys.path.insert(0, str(API_ROOT))

from app.connectors.canadiantire_connector import CanadianTireConnector
from app.connectors.homedepot_connector import HomeDepotConnector
from app.connectors.kms_connector import KMSConnector
from app.connectors.whitecap_connector import WhiteCapConnector
from app.services.scn_catalog_service import SCNBatchIngestService
from app.services.search_service import SearchService

DEFAULT_INPUT_DIR = Path(__file__).resolve().parent / "input"
DEFAULT_SCN_CSV = DEFAULT_INPUT_DIR / "scn_pricing.csv"
DEFAULT_CONTENT_CSV = DEFAULT_INPUT_DIR / "contentlicensing.csv"
DEFAULT_PRICING_XLSX = DEFAULT_INPUT_DIR / "pricing.xlsx"

OUTPUT_COLUMNS = [
    "model",
    "manufacturer_model",
    "description",
    "list_price",
    "distributor_cost",
    "scn_image",
    "unit",
    "manufacturer",
    "warehouse",
]
DEFAULT_SCN_IMAGE_BASE_URL = "https://www.scnindustrial.com/images/xlarge"

LOGGER = logging.getLogger(__name__)


def normalize_key(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower()).strip("_")


def normalize_model(value: object) -> str:
    return str(value or "").strip().upper()


def _sheet_rows_with_headers(worksheet) -> Iterable[dict[str, object]]:
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
    except StopIteration:
        return

    for row in rows:
        yield {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}


def _find_column(fieldnames: list[str], aliases: tuple[str, ...], file_path: Path) -> str:
    by_normalized = {normalize_key(name): name for name in fieldnames if name}
    for alias in aliases:
        if alias in by_normalized:
            return by_normalized[alias]
    raise ValueError(
        f"Could not find any of columns {aliases} in {file_path}. Available columns: {fieldnames}"
    )


def _find_optional_column(fieldnames: list[str], aliases: tuple[str, ...]) -> str | None:
    by_normalized = {normalize_key(name): name for name in fieldnames if name}
    for alias in aliases:
        if alias in by_normalized:
            return by_normalized[alias]
    return None


def _resolve_content_csv(content_csv: Path) -> Path:
    if content_csv.exists():
        return content_csv

    legacy_path = content_csv.with_name("contentLicensingnew.csv")
    if content_csv.name.lower() == "contentlicensing.csv" and legacy_path.exists():
        return legacy_path

    raise FileNotFoundError(f"Content CSV not found: {content_csv}")


def read_content_product_map(content_csv: Path) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    with content_csv.open("r", encoding="utf-8-sig", newline="") as csv_file:
        reader = csv.DictReader(csv_file)
        if not reader.fieldnames:
            raise ValueError(f"Content licensing CSV has no header row: {content_csv}")

        prod_column = _find_column(reader.fieldnames, ("prod",), content_csv)
        brand_column = _find_column(
            reader.fieldnames,
            (
                "brand",
                "marque",
                "manufacturer",
                "fabricant",
                "mfr",
                "mfg",
            ),
            content_csv,
        )
        manufacturer_number_column = _find_column(
            reader.fieldnames,
            ("manufacturernumber", "manufacturer_number", "manufacturer_number_numero_fabricant"),
            content_csv,
        )
        product_title_column = _find_column(
            reader.fieldnames,
            ("producttitle", "product_title", "english_description", "description"),
            content_csv,
        )
        image_main_column = _find_optional_column(
            reader.fieldnames,
            ("image_main", "imagemain", "imagefilename", "image_file", "image"),
        )

        product_by_sku: dict[str, dict[str, str]] = {}
        product_by_manufacturer_model: dict[str, dict[str, str]] = {}
        for row in reader:
            sku = _coerce_text(row.get(prod_column))
            image_main = _coerce_text(row.get(image_main_column)) if image_main_column else ""
            product = {
                "sku": sku,
                "manufacturer": _coerce_text(row.get(brand_column)),
                "manufacturer_model": _coerce_text(row.get(manufacturer_number_column)),
                "title": _coerce_text(row.get(product_title_column)),
                "image_main": image_main,
                "scn_image": _build_scn_image_url(image_main=image_main, prod=sku),
            }
            sku_key = normalize_model(product["sku"])
            manufacturer_model_key = normalize_model(product["manufacturer_model"])

            if not sku_key and not manufacturer_model_key:
                continue

            if sku_key and (
                sku_key not in product_by_sku
                or (not product_by_sku[sku_key].get("manufacturer") and product["manufacturer"])
            ):
                product_by_sku[sku_key] = product

            if manufacturer_model_key and (
                manufacturer_model_key not in product_by_manufacturer_model
                or (
                    not product_by_manufacturer_model[manufacturer_model_key].get("manufacturer")
                    and product["manufacturer"]
                )
            ):
                product_by_manufacturer_model[manufacturer_model_key] = product

    return product_by_sku, product_by_manufacturer_model


def _coerce_text(value: object) -> str:
    return str(value or "").strip()


def _sanitize_path_component(value: str) -> str:
    cleaned = _coerce_text(value).replace("\\", "/")
    cleaned = cleaned.split("/")[-1]
    return re.sub(r"[^A-Za-z0-9._-]+", "", cleaned)


def _build_scn_image_url(*, image_main: str, prod: str) -> str | None:
    sanitized_prod = _sanitize_path_component(prod)
    sanitized_image = _sanitize_path_component(image_main)
    if not sanitized_prod or not sanitized_image:
        return None

    if "." not in Path(sanitized_image).name:
        sanitized_image = f"{sanitized_image}.jpg"

    subdirectory = sanitized_prod[0].lower()
    base_url = DEFAULT_SCN_IMAGE_BASE_URL.rstrip("/")
    return f"{base_url}/{subdirectory}/{sanitized_image}"


def _is_missing(value: object) -> bool:
    if value is None:
        return True
    normalized = str(value).strip().lower()
    return normalized in {"", "null", "none", "nan", "n/a"}


def prepare_csv_for_supabase_ingest(csv_path: Path) -> tuple[Path, dict[str, int]]:
    if not csv_path.exists():
        raise FileNotFoundError(f"SCN CSV not found: {csv_path}")

    filtered_csv_path = csv_path.with_name(f"{csv_path.stem}.filtered_for_ingest{csv_path.suffix}")

    read_rows = 0
    written_rows = 0
    skipped_all_missing = 0
    missing_model_or_manufacturer = 0

    with csv_path.open("r", newline="", encoding="utf-8-sig") as in_handle:
        reader = csv.DictReader(in_handle)
        fieldnames = reader.fieldnames or []
        if not fieldnames:
            raise ValueError(f"CSV has no header row: {csv_path}")

        with filtered_csv_path.open("w", newline="", encoding="utf-8") as out_handle:
            writer = csv.DictWriter(out_handle, fieldnames=fieldnames)
            writer.writeheader()

            for row in reader:
                read_rows += 1
                model = row.get("model")
                manufacturer = row.get("manufacturer")
                warehouse = row.get("warehouse")

                missing_model = _is_missing(model)
                missing_manufacturer = _is_missing(manufacturer)
                missing_warehouse = _is_missing(warehouse)

                if missing_model and missing_manufacturer and missing_warehouse:
                    skipped_all_missing += 1
                    LOGGER.warning(
                        "Skipping CSV row %s (model/manufacturer/warehouse all missing): %s",
                        read_rows,
                        row,
                    )
                    continue

                if missing_model or missing_manufacturer:
                    missing_model_or_manufacturer += 1
                    LOGGER.warning(
                        "CSV row %s missing key field(s): model_missing=%s manufacturer_missing=%s row=%s",
                        read_rows,
                        missing_model,
                        missing_manufacturer,
                        row,
                    )

                writer.writerow(row)
                written_rows += 1

    return filtered_csv_path, {
        "read": read_rows,
        "written": written_rows,
        "skipped_all_missing": skipped_all_missing,
        "missing_model_or_manufacturer": missing_model_or_manufacturer,
    }


def generate_matched_scn_csv(content_csv: Path, pricing_xlsx: Path, output_csv: Path) -> dict[str, int]:
    resolved_content_csv = _resolve_content_csv(content_csv)
    product_by_sku, product_by_manufacturer_model = read_content_product_map(resolved_content_csv)

    workbook = load_workbook(pricing_xlsx, data_only=True, read_only=True)
    output_csv.parent.mkdir(parents=True, exist_ok=True)

    total_rows = 0
    matched_rows = 0
    seen_composite_keys: set[tuple[str, str, str]] = set()

    with output_csv.open("w", encoding="utf-8", newline="") as out_file:
        writer = csv.DictWriter(out_file, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row in _sheet_rows_with_headers(worksheet):
                total_rows += 1
                normalized_row = {normalize_key(k): v for k, v in row.items() if k}
                pricing_model = _coerce_text(normalized_row.get("model_no_no_modele"))
                pricing_mfg_model = _coerce_text(normalized_row.get("mfg_model_no_no_fab"))
                model_key = normalize_model(pricing_model)
                mfg_model_key = normalize_model(pricing_mfg_model)

                matched_product = None
                if model_key:
                    matched_product = product_by_sku.get(model_key)
                elif mfg_model_key:
                    matched_product = product_by_manufacturer_model.get(mfg_model_key)

                if not matched_product:
                    continue

                description = _coerce_text(
                    matched_product.get("title")
                    or normalized_row.get("english_description_description_anglais")
                    or normalized_row.get("description")
                    or matched_product.get("sku")
                )
                list_price = _coerce_text(normalized_row.get("list_price_prix_liste") or normalized_row.get("list_price"))
                distributor_cost = _coerce_text(
                    normalized_row.get("distributor_cost_cout_distributeur")
                    or normalized_row.get("distributor_cost")
                )
                unit = _coerce_text(
                    normalized_row.get("unit_of_sale")
                    or normalized_row.get("unite_de_vente")
                    or normalized_row.get("unit")
                )
                manufacturer = _coerce_text(
                    normalized_row.get("manufacturer_fabricant")
                    or normalized_row.get("manufacturer")
                    or matched_product.get("manufacturer")
                )
                model = _coerce_text(matched_product.get("sku"))
                manufacturer_model = _coerce_text(
                    matched_product.get("manufacturer_model") or pricing_mfg_model
                )
                warehouse = _coerce_text(sheet_name).upper()
                composite_key = (normalize_model(model), manufacturer.strip().upper(), warehouse)
                if composite_key in seen_composite_keys:
                    continue
                seen_composite_keys.add(composite_key)

                writer.writerow(
                    {
                        "model": model,
                        "manufacturer_model": manufacturer_model,
                        "description": description,
                        "list_price": list_price,
                        "distributor_cost": distributor_cost,
                        "scn_image": _coerce_text(matched_product.get("scn_image")),
                        "unit": unit,
                        "manufacturer": manufacturer,
                        "warehouse": warehouse,
                    }
                )
                matched_rows += 1

    return {"processed": total_rows, "matched": matched_rows}


async def ingest_connector_prices(csv_path: Path) -> dict[str, int]:
    search_service = SearchService(
        connectors=[
            WhiteCapConnector(),
            KMSConnector(),
            CanadianTireConnector(),
            HomeDepotConnector(),
        ]
    )
    rows_processed = 0
    rows_with_results = 0
    stored_results = 0

    with csv_path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows_processed += 1
            model = (row.get("model") or "").strip()
            description = (row.get("description") or "").strip()
            query = model or description
            if not query:
                continue

            results, _, _ = await search_service.collect_live_results(query)
            if not results:
                continue

            rows_with_results += 1
            stored_results += len(results)

    return {
        "rows_processed": rows_processed,
        "rows_with_results": rows_with_results,
        "stored_results": stored_results,
    }


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

    parser = argparse.ArgumentParser(
        description=(
            "Generate scn_pricing.csv from pricing/content licensing matched rows and ingest into Supabase"
        )
    )
    parser.add_argument(
        "--csv",
        dest="csv_path",
        default=None,
        help=(
            "Path to SCN CSV output/input. Defaults to SCN_PRICING_CSV or "
            f"{DEFAULT_SCN_CSV}"
        ),
    )
    parser.add_argument(
        "--content-csv",
        default=str(DEFAULT_CONTENT_CSV),
        help=f"Path to content licensing CSV. Default: {DEFAULT_CONTENT_CSV}",
    )
    parser.add_argument(
        "--pricing-xlsx",
        default=str(DEFAULT_PRICING_XLSX),
        help=f"Path to pricing workbook. Default: {DEFAULT_PRICING_XLSX}",
    )
    parser.add_argument(
        "--skip-generate",
        action="store_true",
        help="Skip generating scn_pricing.csv and ingest the selected CSV directly.",
    )
    args = parser.parse_args()

    selected_csv = Path(args.csv_path or os.getenv("SCN_PRICING_CSV") or str(DEFAULT_SCN_CSV))

    if not args.skip_generate:
        content_csv = Path(args.content_csv)
        pricing_xlsx = Path(args.pricing_xlsx)
        content_csv = _resolve_content_csv(content_csv)
        if not pricing_xlsx.exists():
            raise FileNotFoundError(f"Pricing workbook not found: {pricing_xlsx}")

        stats = generate_matched_scn_csv(content_csv, pricing_xlsx, selected_csv)
        print(
            f"Generated {selected_csv} with {stats['matched']} matched rows "
            f"from {stats['processed']} processed pricing rows."
        )

    prepared_csv, prep_stats = prepare_csv_for_supabase_ingest(selected_csv)
    print(
        f"Prepared ingest CSV {prepared_csv} from {prep_stats['read']} rows; "
        f"kept {prep_stats['written']} rows; "
        f"skipped {prep_stats['skipped_all_missing']} rows with model/manufacturer/warehouse all missing; "
        f"logged {prep_stats['missing_model_or_manufacturer']} rows missing model and/or manufacturer."
    )

    service = SCNBatchIngestService()
    result = service.ingest_csv_to_supabase(csv_path=str(prepared_csv))
    print(f"Read {result['read']} rows and upserted {result['upserted']} rows into Supabase.")

    connector_stats = asyncio.run(ingest_connector_prices(prepared_csv))
    print(
        "Connector ingestion completed. "
        f"Processed {connector_stats['rows_processed']} rows, "
        f"stored connector prices for {connector_stats['rows_with_results']} rows, "
        f"inserted {connector_stats['stored_results']} connector result rows."
    )


if __name__ == "__main__":
    main()
